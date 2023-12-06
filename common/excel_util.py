# -*- coding: utf-8 -*-

"""
@author: shencheng
@software: PyCharm
@description: test
@time: 2022/4/13 9:34
"""
import os

import openpyxl

class ExcelUtil:

    def dirname(self):
        return os.path.abspath(os.path.dirname(__file__)).split("common")[0]

    def excel_read(self,filename):
        wb=openpyxl.load_workbook(self.dirname()+r'/data/'+filename+'.xlsx')
        sheet=wb.worksheets[0]
        row=sheet.max_row
        col=sheet.max_column
        row_list=[]
        for r in range(2,row+1):
            col_list=[]
            for col in range(1,col+1):
                datavalue=sheet.cell(row=r,column=col).value
                if datavalue is None:
                    datavalue=''
                    col_list.append(datavalue)
                else:
                    col_list.append(datavalue)
            row_list.append(col_list)
        return row_list

    def excel_read_test(self,filename):
        wb=openpyxl.load_workbook(self.dirname()+r'/data/'+filename+'.xlsx')
        sheetname=wb['测试']
        row_max=sheetname.max_row
        col_max=sheetname.max_column
        iter=sheetname.iter_rows(min_row=2,max_row=row_max,min_col=1,max_col=col_max)
        list=[]
        for rows in iter:
            col_list=[]
            for col in rows:
                datavalue = col.value
                if datavalue is None:
                    datavalue=''
                    col_list.append(datavalue)
                else:
                    col_list.append(datavalue)
            list.append(col_list)
        return list

    def write_excel(self,filename,index,value):
        wb = openpyxl.load_workbook(self.dirname() + r'/data/' + filename + '.xlsx')
        sheetname = wb['测试']
        max_col=sheetname.max_column
        for col in range(1,max_col+1):
            title=sheetname.cell(1,col).value
            if title == '实际结果':
                col_value=col
                sheetname.cell(index+1,col_value,value=value)
        wb.save(self.dirname() + r'/data/' + filename + '.xlsx')



if __name__ == '__main__':
    ExcelUtil().write_excel('login_data_test',1,'测试失败')

